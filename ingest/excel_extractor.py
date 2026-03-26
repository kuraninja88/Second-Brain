"""
Extract structured data from Literature Review.xlsx.

Outputs:
  - Rows inserted into SQLite papers table
  - Text chunks for ChromaDB (one per paper row)
"""
from pathlib import Path
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))
from config import LIT_REVIEW_XLSX


COLUMN_MAP = {
    # Excel column header fragment → our field name
    "no":           "no",
    "date":         "date",
    "author":       "author",
    "reference":    "reference",
    "theme":        "theme",
    "research":     "rq_focus",
    "key theory":   "theory",
    "data":         "data_sample",
    "main var":     "variables",
    "method":       "methodology",
    "key find":     "key_findings",
    "novelty":      "novelty",
    "limitation":   "limitations",
    "notes":        "notes",
    "relevant quote":"relevant_quote",
    "citation":     "citation",
}


def _map_header(header: str) -> str:
    h = str(header).lower().strip()
    for key, field in COLUMN_MAP.items():
        if key in h:
            return field
    return h.replace(" ", "_")[:30]


def _cell(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def extract_excel(xlsx_path: Path = LIT_REVIEW_XLSX) -> list[dict]:
    """
    Returns a list of chunk dicts:
        {
          "text":     <concatenated text for embedding>,
          "metadata": {source_type, file_path, author, year, theme, ...},
          "db_row":   <dict for SQLite upsert>
        }
    """
    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

    results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        # First non-empty row is headers
        headers = [_map_header(h) if h else f"col{i}" for i, h in enumerate(rows[0])]

        for raw_row in rows[1:]:
            row = {headers[i]: _cell(v) for i, v in enumerate(raw_row) if i < len(headers)}

            # Skip completely empty rows
            if not any(row.values()):
                continue

            author = row.get("author", "")
            if not author:
                continue

            # Build readable text for embedding
            parts = []
            if row.get("author"):
                parts.append(f"Author: {row['author']}")
            if row.get("date"):
                parts.append(f"Year: {row['date']}")
            if row.get("reference"):
                parts.append(f"Reference: {row['reference']}")
            if row.get("theme"):
                parts.append(f"Theme: {row['theme']}")
            if row.get("rq_focus"):
                parts.append(f"Research question: {row['rq_focus']}")
            if row.get("theory"):
                parts.append(f"Theory/Framework: {row['theory']}")
            if row.get("data_sample"):
                parts.append(f"Data and sample: {row['data_sample']}")
            if row.get("variables"):
                parts.append(f"Variables: {row['variables']}")
            if row.get("methodology"):
                parts.append(f"Methodology: {row['methodology']}")
            if row.get("key_findings"):
                parts.append(f"Key findings: {row['key_findings']}")
            if row.get("novelty"):
                parts.append(f"Novelty: {row['novelty']}")
            if row.get("limitations"):
                parts.append(f"Limitations/Gap: {row['limitations']}")
            if row.get("notes"):
                parts.append(f"Relevance to thesis: {row['notes']}")
            if row.get("relevant_quote"):
                parts.append(f"Relevant quote: {row['relevant_quote']}")

            text = ". ".join(p for p in parts if p)

            # Extract year from date field (may be "9/9" or "2022" etc.)
            year_raw = row.get("date", "")
            year = year_raw if len(year_raw) == 4 else ""

            metadata = {
                "source_type": "literature_review",
                "file_path":   str(xlsx_path),
                "sheet":       sheet_name,
                "author":      author,
                "year":        year,
                "theme":       row.get("theme", ""),
                "methodology": row.get("methodology", ""),
                "citation":    row.get("citation", ""),
            }

            db_row = {
                "author":        author,
                "year":          year,
                "reference":     row.get("reference", ""),
                "theme":         row.get("theme", ""),
                "rq_focus":      row.get("rq_focus", ""),
                "theory":        row.get("theory", ""),
                "data_sample":   row.get("data_sample", ""),
                "variables":     row.get("variables", ""),
                "methodology":   row.get("methodology", ""),
                "key_findings":  row.get("key_findings", ""),
                "novelty":       row.get("novelty", ""),
                "limitations":   row.get("limitations", ""),
                "notes":         row.get("notes", ""),
                "relevant_quote":row.get("relevant_quote", ""),
                "citation":      row.get("citation", ""),
                "source_file":   str(xlsx_path),
            }

            results.append({"text": text, "metadata": metadata, "db_row": db_row})

    return results
